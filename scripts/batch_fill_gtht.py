#!/usr/bin/env python3
"""
批量用GTHT补齐V7回测Excel中的"无候选"日期
"""
import json, os, sys, subprocess, time, re, math
from openpyxl import load_workbook
from datetime import datetime, timedelta

WORKSPACE = os.path.expanduser('~/.openclaw/workspace')
SKILL_DIR = os.path.join(WORKSPACE, 'skills', 'gtht-smartstockselection-skill')
OUTPUT = os.path.join(WORKSPACE, 'v7_gtht_backtest.xlsx')

# 高危板块
HIGH_RISK = ['房地产','国防军工','食品饮料','非银金融','机械设备','传媒']

# V7 评分权重
W = {'score': 25, 'flow': 20, 'ratio': 15, 'lb': 15, 'hs': 10, 'sector': 15}

def gtht_query(query, retries=2):
    """调用GTHT financial-search，返回解析后的dict列表"""
    cmd = [
        'node', 'skill-entry.js', 'mcpClient', 'call', 'financial', 'financial-search',
        f'query={query}'
    ]
    for attempt in range(retries + 1):
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=30, cwd=SKILL_DIR)
            if r.returncode != 0:
                if attempt < retries:
                    time.sleep(3)
                    continue
                return None
            raw = r.stdout.strip()
            if not raw:
                return None
            # 外层JSON
            obj = json.loads(raw)
            result_str = obj.get('text', '')
            if not result_str:
                return None
            # 内层不是JSON，是markdown表格文本，提取|...|行
            lines = result_str.split('\n')
            # 找表头行（第一个完整|...|行）
            header_line = None
            data_lines = []
            found_sep = False
            for line in lines:
                stripped = line.strip()
                if stripped.startswith('|') and stripped.endswith('|'):
                    if '股票代码' in stripped:
                        header_line = stripped
                    elif '---' in stripped:
                        found_sep = True
                    elif found_sep and header_line:
                        data_lines.append(stripped)
            if not header_line or not data_lines:
                return []
            # 解析表头
            headers = [h.strip() for h in header_line.strip('|').split('|')]
            # 解析数据行
            rows = []
            for dl in data_lines:
                cells = [c.strip() for c in dl.strip('|').split('|')]
                if len(cells) == len(headers):
                    rows.append(dict(zip(headers, cells)))
                elif len(cells) > len(headers):
                    rows.append(dict(zip(headers, cells[:len(headers)])))
                # 跳过列数不匹配的行
            return rows
        except Exception as e:
            if attempt < retries:
                time.sleep(3)
                continue
            print(f"  ⚠️ GTHT异常: {e}")
            return None
    return None

def extract_fields(row, date_str_prefix=''):
    """从GTHT返回行中提取关键字段。优先取带日期后缀的列"""
    try:
        code = row.get('股票代码', '')
        name = row.get('股票简称', '')
        
        suffix = ''
        if date_str_prefix:
            suffix = f'[{date_str_prefix}]'
        
        sector = ''
        zdf = lb = hs = flow = amount = open_p = high_p = close_p = zhenfu = None
        
        # 优先找带日期后缀的列，再fallback
        for k, v in row.items():
            if not v: continue
            # 涨跌幅：优先带日期后缀且不含前复权
            if ('涨跌幅' + suffix) == k or (suffix and '涨跌幅' in k and suffix in k and '前复权' not in k):
                zdf = float(v)
            elif not suffix and '涨跌幅' in k and '前复权' not in k and '最新涨跌幅' != k:
                zdf = float(v)
            # 量比
            if ('量比' + suffix) == k or (suffix and '量比' in k and suffix in k):
                lb = float(v)
            elif not suffix and '量比' in k and '量比' != k:
                lb = float(v)
            # 换手率
            if ('换手率' + suffix) == k or (suffix and '换手率' in k and suffix in k):
                hs = float(v)
            elif not suffix and '换手率' in k and '换手率' != k:
                hs = float(v)
            # 主力资金流向
            if ('主力资金流向' + suffix) == k or (suffix and '主力资金流向' in k and suffix in k):
                flow = float(v)
            elif '主力资金流向' in k and '主力资金流向' != k:
                flow = float(v)
            # 成交额
            if ('成交额' + suffix) == k or (suffix and '成交额' in k and suffix in k):
                amount = float(v)
            elif not suffix and '成交额' in k and '成交额' != k:
                amount = float(v)
            # 开盘价_前复权
            if ('开盘价_前复权' + suffix) == k:
                open_p = float(v)
            elif suffix and '开盘价_前复权' in k and suffix in k:
                open_p = float(v)
            # 最高价_前复权
            if ('最高价_前复权' + suffix) == k:
                high_p = float(v)
            elif suffix and '最高价_前复权' in k and suffix in k:
                high_p = float(v)
            # 收盘价_前复权
            if ('收盘价_前复权' + suffix) == k:
                close_p = float(v)
            elif suffix and '收盘价_前复权' in k and suffix in k:
                close_p = float(v)
            # 振幅
            if ('振幅' + suffix) == k or (suffix and '振幅' in k and suffix in k):
                zhenfu = float(v)
            # 行业（无日期后缀）
            if '所属同花顺行业' in k or '所属行业' in k or ('行业' in k and '所属' not in k and '行业' != k):
                s = str(v).strip('[]').replace("'","").split(',')[0].strip()
                if s: sector = s
        
        # Fallback：如果suffix没匹配到，用最新涨跌幅
        if zdf is None and row.get('最新涨跌幅'):
            zdf = float(row['最新涨跌幅'])
        
        return {
            'code': code,
            'name': name,
            'zdf': zdf,
            'lb': lb,
            'hs': hs,
            'flow': flow if flow else 0,
            'sector': sector,
            'amount': amount if amount else 0,
            'open_p': open_p,
            'high_p': high_p,
            'close_p': close_p,
            'zhenfu': zhenfu,
        }
    except Exception as e:
        # print(f'  extract_fields err: {e}')
        return None

def pn(v):
    if v is None: return 0
    return float(v)

def calc_score(stock):
    """计算V7评分"""
    zdf = pn(stock.get('zdf'))
    lb = pn(stock.get('lb'))
    hs = pn(stock.get('hs'))
    flow = pn(stock.get('flow'))
    amount = pn(stock.get('amount'))
    sector = stock.get('sector', '')
    
    flow_yi = flow / 1e8
    amount_yi = amount / 1e8
    ratio = (flow_yi / amount_yi * 100) if amount_yi > 0 else 0
    
    score = 0
    
    # 净流入评分 (满分20)
    if flow_yi >= 5: score += 20
    elif flow_yi >= 3: score += 16
    elif flow_yi >= 1: score += 12
    elif flow_yi >= 0.5: score += 8
    else: score += 4
    
    # 流入占比评分 (满分15)
    if ratio >= 20: score += 15
    elif ratio >= 10: score += 12
    elif ratio >= 7: score += 9
    elif ratio >= 5: score += 6
    elif ratio >= 3: score += 3
    
    # 量比评分 (满分15)
    if lb >= 3: score += 15
    elif lb >= 2: score += 12
    elif lb >= 1.5: score += 9
    elif lb >= 1.2: score += 6
    
    # 换手评分 (满分10)
    if 5 <= hs <= 10: score += 10
    elif 3 <= hs <= 12: score += 8
    elif 2 <= hs <= 15: score += 5
    
    # 涨幅评分 (满分25)
    if 3 <= zdf <= 6: score += 25
    elif zdf >= 2 and zdf < 3: score += 20
    elif zdf > 6 and zdf <= 7: score += 18
    elif zdf > 7 and zdf <= 8: score += 12
    elif zdf >= 2: score += 10
    
    # 板块协同评分 (满分15) - 在筛选后由外部赋值
    # 这里先留空
    
    return score, flow_yi, ratio

def query_and_screen(date_str):
    """对某一天做V7全筛选"""
    query = f'{date_str} 涨幅大于2%小于8% 量比大于1.2 换手率大于2%小于16% 主力资金净流入大于5000万 非科创板 非创业板'
    rows = gtht_query(query)
    if not rows:
        return None, []
    
    # 提取日期数字前缀，如2026-01-19 -> 20260119
    date_prefix = date_str.replace('-', '')
    candidates = []
    for row in rows:
        stock = extract_fields(row, date_prefix)
        if not stock:
            continue
        
        # 基础条件二次确认
        zdf = pn(stock.get('zdf'))
        lb = pn(stock.get('lb'))
        hs = pn(stock.get('hs'))
        flow = pn(stock.get('flow'))
        amount = pn(stock.get('amount'))
        sector = stock.get('sector', '')
        
        if flow < 5e7:  # 净流入>5000万
            continue
        if hs > 16 or hs < 2:
            continue
        if lb < 1.2:
            continue
        if zdf < 2 or zdf > 8:
            continue
        
        # 排除高危板块（但先保留原始信息，后面再过滤）
        stock['flow_yi'] = flow / 1e8
        stock['amount_yi'] = amount / 1e8
        stock['ratio'] = (stock['flow_yi'] / stock['amount_yi'] * 100) if stock['amount_yi'] > 0 else 0
        
        candidates.append(stock)
    
    if not candidates:
        return None, []
    
    # GTHT financial-search不返回行业字段，无法做板块聚类
    # 历史回填场景：只用核心V7条件，跳过板块≥2/高危排除
    # 只取评分最高的前1只作为当日首推
    valid = candidates
    
    if not valid:
        return None, candidates
    
    # 评分（无板块协同加分，因为GTHT不返回行业字段）
    for c in valid:
        s, f, r = calc_score(c)
        c['v7_score'] = s
        c['flow_yi'] = f
        c['ratio'] = r
    
    valid.sort(key=lambda x: x['v7_score'], reverse=True)
    
    return valid[0], valid  # 首推 + 全部候选

def query_next_day(code, date_str, prev_close=None):
    """查T+1日行情验证，返回百分比涨跌幅"""
    from datetime import datetime, timedelta
    d = datetime.strptime(date_str, '%Y-%m-%d')
    nd = d + timedelta(days=1)
    for _ in range(5):
        if nd.weekday() < 5:
            break
        nd += timedelta(days=1)
    
    next_date = nd.strftime('%Y-%m-%d')
    nprefix = next_date.replace('-', '')
    query = f'{next_date} {code} 涨跌幅 开盘价_前复权 最高价_前复权 收盘价_前复权'
    rows = gtht_query(query)
    if not rows:
        return None
    
    for row in rows:
        stock_code = row.get('股票代码', '')
        if stock_code.split('.')[0] != code.split('.')[0]:
            continue
        
        ns = f'[{nprefix}]'
        
        # T+1日行情
        n_open_price = row.get(f'开盘价_前复权{ns}', '')
        n_high_price = row.get(f'最高价_前复权{ns}', '')
        n_close_price = row.get(f'收盘价_前复权{ns}', '')
        
        # 涨跌幅[YYYYMMDD] 就是收盘涨跌幅（%），直接是next_close
        close_zdf_raw = row.get(f'涨跌幅{ns}', '')
        close_zdf = float(close_zdf_raw) if close_zdf_raw else None
        
        # 计算open和high的涨跌幅%
        # prev_close = close_price / (1 + close_zdf/100)
        # open% = (open_price - prev_close) / prev_close * 100
        open_zdf = None
        high_zdf = None
        
        if n_close_price and close_zdf is not None:
            prev_c = float(n_close_price) / (1 + close_zdf / 100)
            if prev_c > 0:
                if n_open_price:
                    open_zdf = round((float(n_open_price) - prev_c) / prev_c * 100, 2)
                if n_high_price:
                    high_zdf = round((float(n_high_price) - prev_c) / prev_c * 100, 2)
        
        return {
            'next_open': open_zdf,
            'next_high': high_zdf,
            'next_close': close_zdf,
            'next_date': next_date
        }
    return None

def update_excel(ws, row_num, stock, next_day):
    """写入Excel行"""
    # 状态
    ws.cell(row_num, 3, 'OK')
    # 代码
    ws.cell(row_num, 4, stock['code'])
    # 名称
    ws.cell(row_num, 5, stock['name'])
    # 行业
    ws.cell(row_num, 6, stock.get('sector', ''))
    # 涨幅
    ws.cell(row_num, 7, round(stock['zdf'], 2) if stock.get('zdf') else None)
    # 量比
    ws.cell(row_num, 8, round(stock['lb'], 2) if stock.get('lb') else None)
    # 换手
    ws.cell(row_num, 9, round(stock['hs'], 2) if stock.get('hs') else None)
    # 净流入(亿)
    ws.cell(row_num, 10, round(stock.get('flow_yi', 0), 2))
    # 流入占比
    ws.cell(row_num, 11, round(stock.get('ratio', 0), 2))
    # 评分
    ws.cell(row_num, 12, stock.get('v7_score', 0))
    # 当日OHLC略
    
    if next_day:
        ws.cell(row_num, 17, round(next_day['next_open'], 2) if next_day.get('next_open') else None)
        ws.cell(row_num, 18, round(next_day['next_high'], 2) if next_day.get('next_high') else None)
        ws.cell(row_num, 19, round(next_day['next_close'], 2) if next_day.get('next_close') else None)
        
        # 判定
        high = next_day.get('next_high')
        close = next_day.get('next_close')
        if high is not None:
            if high >= 2:
                ws.cell(row_num, 21, '✅达标')
            else:
                ws.cell(row_num, 21, '❌失败')
        else:
            ws.cell(row_num, 21, '待验证')

def get_trading_days():
    """获取需要补齐的交易日"""
    wb = load_workbook(OUTPUT)
    ws = wb.active
    
    empty_dates = []
    for r in range(3, ws.max_row + 1):
        date = ws.cell(r, 1).value
        status = ws.cell(r, 3).value
        if date and isinstance(date, str) and date.startswith('20'):
            status_val = (status or '').strip()
            if status_val == '无候选' or not status_val:
                empty_dates.append((r, date))
    
    wb.close()
    return empty_dates

def main():
    # Check auth
    entry = os.path.join(WORKSPACE, 'skills', 'gtht-skill-shared', 'gtht-entry.json')
    if not os.path.exists(entry):
        print("❌ GTHT授权文件不存在")
        sys.exit(1)
    print(f"✅ GTHT已授权")
    
    empty_dates = get_trading_days()
    print(f"📊 待补齐: {len(empty_dates)} 天")
    
    wb = load_workbook(OUTPUT)
    ws = wb.active
    
    quota = 0
    MAX_QUOTA = 50  # 本次最多用50次（之后手动续跑）
    
    for idx, (row_num, date_str) in enumerate(empty_dates):
        if quota >= MAX_QUOTA:
            print(f"\n⏸️ 已达本批配额 {MAX_QUOTA}次，剩余 {len(empty_dates) - idx} 天待下批")
            break
        
        print(f"\n[{idx+1}/{len(empty_dates)}] {date_str} R{row_num}...", end=' ', flush=True)
        
        # Step 1: V7筛选
        top_pick, all_cands = query_and_screen(date_str)
        quota += 1
        
        if top_pick:
            print(f"首推 {top_pick['name']}({top_pick['code']}) 评分{top_pick.get('v7_score',0)} {top_pick.get('sector','')}", end='')
            
            # Step 2: 查下个交易日验证
            next_day = query_next_day(top_pick['code'], date_str)
            quota += 1
            
            if next_day:
                print(f" → 次高{next_day.get('next_high','?')}%", end='')
            
            # Step 3: 写入Excel
            update_excel(ws, row_num, top_pick, next_day)
            
            # 保存
            print(" ✅")
        else:
            # 无候选 - 保持原样
            all_count = len(all_cands) if all_cands else 0
            print(f"无候选 (基础筛选{all_count}只)", end='')
            if all_cands:
                # 记录基础筛选数用于诊断
                pass
            print("")
        
        # 保存
        wb.save(OUTPUT)
        
        # GTHT限速
        time.sleep(1.5)
    
    wb.save(OUTPUT)
    print(f"\n✅ 本批完成: {min(quota, MAX_QUOTA)} 次查询, 更新了 {min(len(empty_dates), idx+1 if quota >= MAX_QUOTA else len(empty_dates))} 天")
    print(f"   下一批命令: python3 scripts/batch_fill_gtht.py")

if __name__ == '__main__':
    main()
